import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx', data_only=False)
ws = wb['ENG']
print('Columns:', [cell.value for cell in ws[1]])
print('First 20 rows:')
for i in range(1, 21):
    print(i, [cell.value for cell in ws[i]])
