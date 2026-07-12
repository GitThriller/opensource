import csv
import openpyxl
from pathlib import Path

workbook_path = Path(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx')
tsv_path = Path(r'c:\Users\erikg\OneDrive\Git\opensource\Common code\python\video_counts.tsv')

counts = {}
with open(tsv_path, 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f, delimiter='\t')
    for row in reader:
        name = row['Folder_Name'].strip()
        counts[name.casefold()] = int(row['Count'])

wb = openpyxl.load_workbook(workbook_path)
ws = wb['ENG']

updates = []
for row in range(2, ws.max_row + 1):
    name_cell = ws[f'A{row}']
    if name_cell.value is None:
        continue
    name_key = str(name_cell.value).casefold()
    expected = counts.get(name_key, 0)
    cell = ws[f'B{row}']
    old = cell.value
    if old != expected:
        updates.append((row, name_cell.value, old, expected))
        cell.value = expected

print('Found', len(updates), 'updates:')
for r, name, old, expected in updates[:50]:
    print(r, name, old, '->', expected)

wb.save(workbook_path)
print('Saved workbook.')
