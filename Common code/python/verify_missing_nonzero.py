import csv
import openpyxl
from pathlib import Path

workbook_path = Path(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx')
tsv_path = Path(r'c:\Users\erikg\OneDrive\Git\opensource\Common code\python\video_counts.tsv')

counts = {}
with open(tsv_path, 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f, delimiter='\t')
    for row in reader:
        counts[row['Folder_Name'].strip().casefold()] = int(row['Count'])

wb = openpyxl.load_workbook(workbook_path, data_only=True)
ws = wb['ENG']
nonzero = []
for row in range(2, ws.max_row + 1):
    name = ws[f'A{row}'].value
    if not name:
        continue
    key = str(name).casefold()
    if key not in counts:
        count = ws[f'B{row}'].value
        if count not in (None, 0, '0'):
            nonzero.append((row, name, count))

print('Unmatched worksheet names with nonzero count:', len(nonzero))
for item in nonzero[:50]:
    print(item)
