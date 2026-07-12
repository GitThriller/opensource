import csv
import openpyxl
from pathlib import Path

workbook_path = Path(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx')
tsv_path = Path(r'c:\Users\erikg\OneDrive\Git\opensource\Common code\python\video_counts.tsv')

counts = {}
with open(tsv_path, 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f, delimiter='\t')
    for row in reader:
        name = row['Folder_Name'].strip()
        counts[name.casefold()] = int(row['Count'])

print('count keys', len(counts))
print('alexa in counts', counts.get('alexa chains'))

wb = openpyxl.load_workbook(workbook_path)
ws = wb['ENG']

row=28
name=ws[f'A{row}'].value
old=ws[f'B{row}'].value
new=counts.get(str(name).casefold(), 0)
print('row',row,'name',repr(name),'old',old,'new',new)
print('old==new',old==new)
print('type old',type(old),'type new',type(new))

for r in [28,41,47,187,190]:
    name=ws[f'A{r}'].value
    old=ws[f'B{r}'].value
    new=counts.get(str(name).casefold(), 0)
    print(r, repr(name), old, new, old==new)
