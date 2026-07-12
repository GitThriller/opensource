import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx')
ws = wb['ENG']

# Get the max row with data
max_row = ws.max_row

# Dynamic formula that checks entire list and expands as new rows are added
dynamic_formula = '=IF(COUNTIF(INDIRECT("A$2:A"&COUNTA($A:$A)),A{})>1,"Duplicate in ENG","")'

# Apply the formula to all data rows (starting from row 2)
for row in range(2, max_row + 1):
    formula = dynamic_formula.format(row)
    ws[f'C{row}'].value = formula

wb.save(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx')
print(f"Applied dynamic formula to rows 2-{max_row}")
print("\nFormula now automatically checks entire column A and expands as you add new rows!")
