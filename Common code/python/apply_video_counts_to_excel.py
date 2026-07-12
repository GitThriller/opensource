import csv
import openpyxl
from pathlib import Path

workbook_path = Path(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx')
tsv_path = Path(r'c:\Users\erikg\OneDrive\Git\opensource\Common code\python\video_counts.tsv')

wb = openpyxl.load_workbook(workbook_path)
ws = wb['ENG']

counts = {}
with open(tsv_path, 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f, delimiter='\t')
    for row in reader:
        name = row['Folder_Name'].strip()
        try:
            count = int(row['Count'])
        except ValueError:
            count = 0
        counts[name.casefold()] = count

worksheet_names = {}
for row in range(2, ws.max_row + 1):
    cell = ws[f'A{row}']
    if cell.value is None:
        continue
    worksheet_names[cell.value.casefold()] = row

updated = 0
missing_in_tsv = []
for name_casefold, row in worksheet_names.items():
    name = ws[f'A{row}'].value
    new_count = counts.get(name_casefold, 0)
    old_count = ws[f'B{row}'].value
    if old_count != new_count:
        ws[f'B{row}'].value = new_count
        updated += 1
    if name_casefold not in counts:
        missing_in_tsv.append(name)

extra_in_tsv = [name for name in counts.keys() if name not in worksheet_names]

wb.save(workbook_path)

print(f'Updated {updated} rows in ENG worksheet.')
print(f'Worksheet names not found in TSV: {len(missing_in_tsv)}')
print(f'TSV names not found in worksheet: {len(extra_in_tsv)}')
if missing_in_tsv:
    print('\nFirst 20 worksheet names missing in TSV:')
    for name in missing_in_tsv[:20]:
        print('  ', name)
if extra_in_tsv:
    print('\nFirst 20 TSV names missing in worksheet:')
    for name in extra_in_tsv[:20]:
        print('  ', name)
