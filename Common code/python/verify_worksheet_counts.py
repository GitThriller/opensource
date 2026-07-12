import csv
import openpyxl
from pathlib import Path

workbook_path = Path(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx')
tsv_path = Path(r'c:\Users\erikg\OneDrive\Git\opensource\Common code\python\video_counts.tsv')

counts = {}
with open(tsv_path, 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f, delimiter='\t')
    for row in reader:
        name = row['Folder_Name'].strip()
        counts[name.casefold()] = int(row['Count'])

wb = openpyxl.load_workbook(workbook_path, data_only=True)
ws = wb['ENG']

mismatches = []
missing_in_tsv = []
for row in range(2, ws.max_row + 1):
    name = ws[f'A{row}'].value
    if name is None:
        continue
    name_key = str(name).casefold()
    count = ws[f'B{row}'].value
    expected = counts.get(name_key, 0)
    if count != expected:
        mismatches.append((row, name, count, expected))
    if name_key not in counts:
        missing_in_tsv.append(name)

extra_in_tsv = [name for name in counts if name not in {ws[f'A{r}'].value.casefold() for r in range(2, ws.max_row+1) if ws[f'A{r}'].value}]

print('Total rows checked:', ws.max_row - 1)
print('Mismatches:', len(mismatches))
if mismatches:
    print('\nFirst 30 mismatches:')
    for item in mismatches[:30]:
        print(item)
print('\nWorksheet names missing in TSV:', len(missing_in_tsv))
print('First 30 missing names:')
for name in missing_in_tsv[:30]:
    print('  ', name)
print('\nTTY counts entries not in worksheet:', len(extra_in_tsv))
for name in extra_in_tsv[:30]:
    print('  ', name)
