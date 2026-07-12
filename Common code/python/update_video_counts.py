import os
import openpyxl
from collections import defaultdict

# Count videos in both folders
def count_videos_by_pattern(folder_path):
    """Count videos, grouped by pattern (removing numbers, brackets, etc.)"""
    if not os.path.isdir(folder_path):
        print(f"Folder not found: {folder_path}")
        return {}
    
    import re
    video_counts = defaultdict(int)
    
    for root, dirs, files in os.walk(folder_path):
        for file_name in files:
            if file_name.lower().endswith(('.mp4', '.avi', '.mkv', '.flv', '.wmv', '.ts', '.mov')):
                # Process the name: remove brackets, parentheses, numbers
                name = os.path.splitext(file_name)[0]
                if name.startswith('['):
                    name = name
                else:
                    name = re.sub(r'\[.*\]', '', name)
                    name = re.sub(r'\(.*\)', '', name)
                    name = re.sub(r'\d{1,}', '', name)
                    name = name.strip()
                
                video_counts[name] += 1
    
    return dict(video_counts)

print("Counting videos in H:\\A\\ENG...")
eng_counts = count_videos_by_pattern(r'H:\A\ENG')
print(f"Found {len(eng_counts)} unique names in ENG folder\n")

print("Counting videos in H:\\A\\Non-ENG...")
non_eng_counts = count_videos_by_pattern(r'H:\A\Non-ENG')
print(f"Found {len(non_eng_counts)} unique names in Non-ENG folder\n")

# Load the Excel file
wb = openpyxl.load_workbook(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx')
ws = wb['ENG']

# Update counts in column B based on actual folder counts
print("Updating counts in ENG worksheet...\n")
updated = 0
mismatches = []

for row in range(2, ws.max_row + 1):
    name = ws[f'A{row}'].value
    current_count = ws[f'B{row}'].value
    
    if name:
        # Look up the count in the eng_counts
        actual_count = eng_counts.get(name, 0)
        
        # Update if different
        if actual_count != current_count:
            ws[f'B{row}'].value = actual_count
            updated += 1
            mismatches.append(f"Row {row}: '{name}' - was {current_count}, now {actual_count}")

# Save the file
wb.save(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx')

print(f"Updated {updated} rows with correct counts\n")

if mismatches:
    print("Mismatches found:")
    for mismatch in mismatches[:20]:  # Show first 20
        print(f"  {mismatch}")
    if len(mismatches) > 20:
        print(f"  ... and {len(mismatches) - 20} more")
else:
    print("All counts match!")
