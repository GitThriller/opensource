import csv
import openpyxl
from pathlib import Path

workbook_path = Path(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx')
tsv_path = Path(r'c:\Users\erikg\OneDrive\Git\opensource\Common code\python\video_counts.tsv')

counts = {}
with open(tsv_path, 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f, delimiter='\t')
    for row in reader:
        name = row['Folder_Name'].strip()
        counts[name.casefold()] = int(row['Count'])

print('count keys', len(counts))
print('alexa in counts', counts.get('alexa chains'))

wb = openpyxl.load_workbook(workbook_path)
ws = wb['ENG']
print('row 28', ws['A28'].value, ws['B28'].value)
print('matching key', ws['A28'].value.casefold())
print('counts get', counts.get(ws['A28'].value.casefold()))
