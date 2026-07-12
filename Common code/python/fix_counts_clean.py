import os
import openpyxl

# Count videos by folder name (excluding ._ files)
def count_videos_by_folder(folder_path):
    """Count videos in each named folder"""
    video_counts = {}
    
    if not os.path.isdir(folder_path):
        print(f"Folder not found: {folder_path}")
        return {}
    
    for item in os.listdir(folder_path):
        item_path = os.path.join(folder_path, item)
        if os.path.isdir(item_path):
            try:
                video_count = 0
                for file in os.listdir(item_path):
                    # Skip ._ buffering files AND desktop.ini
                    if not file.startswith('._') and file != 'desktop.ini' and file.lower().endswith(('.mp4', '.avi', '.mkv', '.flv', '.wmv', '.ts', '.mov')):
                        video_count += 1
                
                if video_count > 0:
                    video_counts[item] = video_count
            except PermissionError:
                pass
    
    return video_counts

print("Counting videos by folder in H:\\A\\ENG (excluding ._)...")
eng_counts = count_videos_by_folder(r'H:\A\ENG')
print(f"Found {len(eng_counts)} named folders\n")

# Show sample counts
print("Sample counts:")
for name, count in list(eng_counts.items())[:10]:
    print(f"  {name}: {count}")

# Load worksheet
wb = openpyxl.load_workbook(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx')
ws = wb['ENG']

# Clear all counts first, then update with new counts
print(f"\nClearing and updating all counts...\n")
updated = 0

for row in range(2, ws.max_row + 1):
    name = ws[f'A{row}'].value
    
    if name:
        actual_count = eng_counts.get(name, 0)
        ws[f'B{row}'].value = actual_count
        updated += 1

# Save
wb.save(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx')

print(f"Updated all {updated} rows with correct counts (._files excluded)")

# Verify
print("\nVerification:")
for name in ['Alexa Chains', 'Becky Bandini', 'Charlotte Sins']:
    if name in eng_counts:
        print(f"  {name}: {eng_counts[name]} videos")
