import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx')
ws = wb['ENG']

print("Sample names from worksheet column A:\n")
for i in range(2, 22):
    name = ws[f'A{i}'].value
    count = ws[f'B{i}'].value
    print(f"Row {i}: '{name}' - Count: {count}")
