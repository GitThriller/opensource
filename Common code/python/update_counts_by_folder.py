import os
import openpyxl
from collections import defaultdict

VIDEO_EXTENSIONS = ('.mp4', '.avi', '.mkv', '.flv', '.wmv', '.ts', '.mov')


def is_video_file(filename: str) -> bool:
    return (
        filename
        and not filename.startswith('._')
        and filename.lower().endswith(VIDEO_EXTENSIONS)
    )


def extract_name_from_filename(filename: str) -> str:
    name = os.path.splitext(filename)[0]
    if '[' in name:
        name = name.split('[', 1)[0]
    return name.strip()


def count_videos_in_directory(root_path: str) -> dict:
    """Count videos by folder name and by loose file name."""
    counts = defaultdict(int)

    if not os.path.isdir(root_path):
        print(f"Folder not found: {root_path}")
        return counts

    for entry in os.listdir(root_path):
        entry_path = os.path.join(root_path, entry)
        if os.path.isdir(entry_path):
            try:
                for filename in os.listdir(entry_path):
                    if is_video_file(filename):
                        counts[entry] += 1
            except PermissionError:
                pass  # Skip folders without access
        elif os.path.isfile(entry_path) and is_video_file(entry):
            name = extract_name_from_filename(entry)
            if name:
                counts[name] += 1

    return counts


def merge_counts(*count_dicts):
    merged = defaultdict(int)
    for d in count_dicts:
        for name, count in d.items():
            merged[name] += count
    return merged


print("Counting videos in H:\\A\\ENG and H:\\A\\Non-ENG...")
eng_counts = count_videos_in_directory(r'H:\A\ENG')
non_eng_counts = count_videos_in_directory(r'H:\A\Non-ENG')
all_counts = merge_counts(eng_counts, non_eng_counts)
print(f"Found {len(all_counts)} unique names across ENG and Non-ENG\n")

# Load the Excel file
wb = openpyxl.load_workbook(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx')
ws = wb['ENG']

# Normalize lookup for case-insensitive matching
normalized_counts = {name.casefold(): count for name, count in all_counts.items()}

print("Updating counts in ENG worksheet...\n")
updated = 0
mismatches = []

for row in range(2, ws.max_row + 1):
    name = ws[f'A{row}'].value
    current_count = ws[f'B{row}'].value
    if name:
        actual_count = normalized_counts.get(str(name).casefold(), 0)
        if actual_count != current_count:
            ws[f'B{row}'].value = actual_count
            updated += 1
            mismatches.append((row, name, current_count, actual_count))

wb.save(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx')

print(f"Updated {updated} rows\n")

if mismatches:
    print(f"Changes made (showing first 30):")
    for row, name, old_count, new_count in mismatches[:30]:
        print(f"  Row {row}: '{name}' - was {old_count}, now {new_count}")
    if len(mismatches) > 30:
        print(f"  ... and {len(mismatches) - 30} more")

print("\nDone! Counts updated in Entertainment.xlsx")
