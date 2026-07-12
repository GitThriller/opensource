import os
import openpyxl

# Check what files are in one of the folders
test_folder = r'H:\A\ENG\Alexa Chains'
print(f"Files in '{test_folder}':\n")
all_files = os.listdir(test_folder)
for f in all_files:
    print(f"  {f}")

print(f"\nTotal files: {len(all_files)}")

# Count with filter
video_count = 0
for f in all_files:
    if not f.startswith('._') and f.lower().endswith(('.mp4', '.avi', '.mkv', '.flv', '.wmv', '.ts', '.mov')):
        video_count += 1

print(f"Videos (excluding ._): {video_count}")

# Check worksheet
wb = openpyxl.load_workbook(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx')
ws = wb['ENG']

# Find Alexa Chains
for row in range(2, ws.max_row + 1):
    if ws[f'A{row}'].value == 'Alexa Chains':
        count = ws[f'B{row}'].value
        print(f"\nWorksheet shows: Alexa Chains - {count} videos")
        break
