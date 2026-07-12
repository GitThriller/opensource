import os
import re
import openpyxl
from pathlib import Path
from collections import defaultdict

VIDEO_EXTENSIONS = ('.mp4', '.avi', '.mkv', '.flv', '.wmv', '.ts', '.mov')


def is_video_file(filename: str) -> bool:
    return (
        filename
        and not filename.startswith('._')
        and filename.lower().endswith(VIDEO_EXTENSIONS)
    )


def normalize_video_name(filename: str) -> str:
    name = os.path.splitext(filename)[0].strip()
    if not name:
        return ''
    if name.startswith('[') and name.endswith(']'):
        return name.strip()
    name = re.sub(r'\[.*?\]', '', name)
    name = re.sub(r'\(.*?\)', '', name)
    name = re.sub(r'\d+', '', name)
    return name.strip()


def count_videos_in_root(root_path: Path) -> dict:
    counts = defaultdict(set)
    if not root_path.is_dir():
        print(f"Folder not found: {root_path}")
        return counts

    for entry in os.listdir(root_path):
        entry_path = root_path / entry
        try:
            is_dir = entry_path.is_dir()
        except PermissionError:
            print(f"Skipping inaccessible entry: {entry_path}")
            continue

        if is_dir:
            try:
                for filename in os.listdir(entry_path):
                    if is_video_file(filename):
                        normalized = normalize_video_name(filename)
                        if normalized:
                            counts[entry.strip().casefold()].add(normalized)
            except PermissionError:
                print(f"Skipping inaccessible folder: {entry_path}")
        else:
            try:
                if entry_path.is_file() and is_video_file(entry):
                    normalized = normalize_video_name(entry)
                    if normalized:
                        counts[normalized.casefold()].add(normalized)
            except PermissionError:
                print(f"Skipping inaccessible file: {entry_path}")

    return counts


def merge_counts(*counts_dicts):
    merged = defaultdict(set)
    for counts in counts_dicts:
        for name, values in counts.items():
            merged[name].update(values)
    return merged


def load_excel_counts(excel_path: Path, sheet_name: str = 'ENG') -> tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]
    return wb, ws


def apply_counts_to_worksheet(ws, counts):
    updated = 0
    missing = []
    for row in range(2, ws.max_row + 1):
        name_cell = ws[f'A{row}']
        if name_cell.value is None:
            continue
        key = str(name_cell.value).strip().casefold()
        new_count = len(counts.get(key, set()))
        old_count = ws[f'B{row}'].value
        if old_count != new_count:
            ws[f'B{row}'].value = new_count
            updated += 1
        if key not in counts:
            missing.append(name_cell.value)
    return updated, missing


def main():
    eng_root = Path(r'H:\A\ENG')
    non_eng_root = Path(r'H:\A\Non-ENG')
    excel_path = Path(r'C:\Users\erikg\OneDrive\Documents\zzz\Entertainment.xlsx')

    eng_counts = count_videos_in_root(eng_root)
    non_eng_counts = count_videos_in_root(non_eng_root)
    combined = merge_counts(eng_counts, non_eng_counts)

    wb, ws = load_excel_counts(excel_path, 'ENG')
    updated, missing = apply_counts_to_worksheet(ws, combined)
    wb.save(excel_path)

    print(f'Updated {updated} rows in ENG worksheet.')
    print(f'Worksheet names not found in any folder count: {len(missing)}')
    if missing:
        print('\nFirst 30 missing names:')
        for name in missing[:30]:
            print('  ', name)

    print('\nSample counts from combined scan:')
    for name in ['Alexa Chains', 'Becky Bandini', 'Charlotte Sins', 'Tiffany Brookes']:
        key = name.casefold()
        print('  ', name, len(combined.get(key, set())))


if __name__ == '__main__':
    main()
